# export.py
# pip install openpyxl reportlab

import io
from datetime import datetime
from flask import Blueprint, send_file
import sqlite3

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import DoughnutChart, Reference

export_bp = Blueprint('export', __name__)

DATABASE = 'elecciones.db'

# ── Paleta ───────────────────────────────────────────────────
ROJO     = '#D7263D'
AZUL     = '#1B4F9B'
AMARILLO = '#F5A623'

RL_AZUL     = colors.HexColor(AZUL)
RL_ROJO     = colors.HexColor(ROJO)
RL_AMARILLO = colors.HexColor(AMARILLO)
RL_GRIS     = colors.HexColor('#F4F6FB')
RL_BLANCO   = colors.white


def _fetch_data():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    rows = db.execute(
        "SELECT name, empleados, votos_reportados FROM secretarias ORDER BY name"
    ).fetchall()
    db.close()
    return [dict(r) for r in rows]


# ════════════════════════════════════════════════════════════
#  EXCEL
# ════════════════════════════════════════════════════════════
@export_bp.route('/export/excel')
def export_excel():
    data      = _fetch_data()
    total_emp = sum(r['empleados']        for r in data)
    total_vot = sum(r['votos_reportados'] for r in data)
    total_falt= total_emp - total_vot

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen General"
    ws.sheet_view.showGridLines = False

    fill_azul     = PatternFill("solid", fgColor=AZUL[1:])
    fill_rojo     = PatternFill("solid", fgColor=ROJO[1:])
    fill_amarillo = PatternFill("solid", fgColor=AMARILLO[1:])
    fill_blanco   = PatternFill("solid", fgColor="FFFFFF")
    fill_gris     = PatternFill("solid", fgColor="F4F6FB")
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    # Título
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = "CONSULTA POPULAR NACIONAL 2026"
    c.font  = Font(bold=True, size=18, color="FFFFFF", name="Calibri")
    c.fill  = fill_azul
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 38

    ws.merge_cells('A2:G2')
    c2 = ws['A2']
    c2.value = "GOBERNACIÓN DEL ESTADO BOLÍVAR — REPORTE DE PARTICIPACIÓN"
    c2.font  = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    c2.fill  = fill_rojo
    c2.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    ws.merge_cells('A3:G3')
    c3 = ws['A3']
    c3.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c3.font  = Font(italic=True, size=9, color="888888", name="Calibri")
    c3.alignment = Alignment(horizontal='right')
    ws.row_dimensions[3].height = 16

    # Encabezados tabla (fila 5)
    headers   = ['#', 'SECRETARÍA', 'EMPLEADOS', 'YA VOTARON', 'FALTAN', '% PARTICIPACIÓN', 'ESTADO']
    col_widths= [4, 54, 13, 13, 13, 18, 14]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font      = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        cell.fill      = fill_azul
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[5].height = 28

    # Filas de datos
    for idx, row in enumerate(data, 1):
        r    = 5 + idx
        falt = row['empleados'] - row['votos_reportados']
        pct  = (row['votos_reportados'] / row['empleados'] * 100) if row['empleados'] > 0 else 0
        estado = "✔ Completo" if falt == 0 else ("⚠ Parcial" if pct > 0 else "✗ Sin votos")
        rfill = fill_gris if idx % 2 == 0 else fill_blanco

        vals   = [idx, row['name'], row['empleados'], row['votos_reportados'], falt, f"{pct:.1f}%", estado]
        aligns = ['center','left','center','center','center','center','center']
        for c, (v, al) in enumerate(zip(vals, aligns), 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill      = rfill
            cell.alignment = Alignment(horizontal=al, vertical='center', wrap_text=(c==2))
            cell.border    = border
            cell.font      = Font(size=9, name="Calibri")
            if c == 6:
                pv = float(str(v).replace('%',''))
                cell.font = Font(bold=True, size=9, name="Calibri",
                                 color="27AE60" if pv>=80 else ("F5A623" if pv>=50 else "D7263D"))
        ws.row_dimensions[r].height = 20

    # Fila total
    tr = 5 + len(data) + 1
    ws.merge_cells(f'A{tr}:B{tr}')
    for col in range(1, 8):
        cell = ws.cell(tr, col)
        cell.fill   = fill_amarillo
        cell.font   = Font(bold=True, size=10, name="Calibri")
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(tr, 1).value = "TOTAL GENERAL"
    ws.cell(tr, 3).value = total_emp
    ws.cell(tr, 4).value = total_vot
    ws.cell(tr, 5).value = total_falt
    ws.cell(tr, 6).value = f"{(total_vot/total_emp*100):.1f}%" if total_emp else "0%"
    ws.row_dimensions[tr].height = 24
    ws.freeze_panes = 'A6'

    # ── Hoja 2: Gráfico ──────────────────────────────────────
    ws2 = wb.create_sheet("Gráfico")
    ws2.sheet_view.showGridLines = False
    ws2['A1'] = "Categoría"; ws2['B1'] = "Cantidad"
    ws2['A2'] = "Ya Votaron"; ws2['B2'] = total_vot
    ws2['A3'] = "Faltan";     ws2['B3'] = total_falt

    chart = DoughnutChart()
    chart.title    = "Participación General"
    chart.style    = 10
    chart.holeSize = 40
    data_ref   = Reference(ws2, min_col=2, min_row=1, max_row=3)
    labels_ref = Reference(ws2, min_col=1, min_row=2, max_row=3)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    chart.width  = 16
    chart.height = 12
    ws2.add_chart(chart, "D2")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"participacion_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ════════════════════════════════════════════════════════════
#  PDF
# ════════════════════════════════════════════════════════════
@export_bp.route('/export/pdf')
def export_pdf():
    data      = _fetch_data()
    total_emp = sum(r['empleados']        for r in data)
    total_vot = sum(r['votos_reportados'] for r in data)
    total_falt= total_emp - total_vot
    pct_gen   = (total_vot / total_emp * 100) if total_emp > 0 else 0

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    title_s = ps('T',  fontSize=20, fontName='Helvetica-Bold',
                 textColor=RL_BLANCO, alignment=TA_CENTER)
    sub_s   = ps('S',  fontSize=10, fontName='Helvetica',
                 textColor=RL_BLANCO, alignment=TA_CENTER)
    normal  = ps('N',  fontSize=8,  fontName='Helvetica')
    bold_c  = ps('BC', fontSize=8,  fontName='Helvetica-Bold', alignment=TA_CENTER)
    pct_s   = ps('P',  fontSize=22, fontName='Helvetica-Bold',
                 textColor=RL_AMARILLO, alignment=TA_CENTER)
    lbl_s   = ps('L',  fontSize=8,  fontName='Helvetica',
                 textColor=colors.HexColor('#555555'), alignment=TA_CENTER)

    elems = []

    # Encabezado
    hdr = Table([
        [Paragraph("CONSULTA POPULAR NACIONAL 2026", title_s)],
        [Paragraph("Gobernación del Estado Bolívar · Reporte de Participación", sub_s)],
        [Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_s)],
    ], colWidths=[doc.width])
    hdr.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,0), RL_AZUL),
        ('BACKGROUND',    (0,1),(-1,2), RL_ROJO),
        ('TOPPADDING',    (0,0),(-1,-1), 8),
        ('BOTTOMPADDING', (0,0),(-1,-1), 8),
    ]))
    elems += [hdr, Spacer(1, .4*cm)]

    # Métricas
    met = Table([[
        Paragraph(f"{total_emp}",       pct_s),
        Paragraph(f"{total_vot}",       pct_s),
        Paragraph(f"{total_falt}",      pct_s),
        Paragraph(f"{pct_gen:.1f}%",    pct_s),
    ],[
        Paragraph("Total Empleados",    lbl_s),
        Paragraph("Ya Votaron",         lbl_s),
        Paragraph("Faltan",             lbl_s),
        Paragraph("% Participación",    lbl_s),
    ]], colWidths=[doc.width/4]*4)
    met.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(0,-1), colors.HexColor('#E8EEF9')),
        ('BACKGROUND', (1,0),(2,-1), colors.HexColor('#FDE8EB')),
        ('BACKGROUND', (3,0),(3,-1), colors.HexColor('#FFF8E7')),
        ('BOX',        (0,0),(-1,-1), 1.5, RL_AZUL),
        ('INNERGRID',  (0,0),(-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING',    (0,0),(-1,-1), 6),
        ('BOTTOMPADDING', (0,0),(-1,-1), 6),
    ]))
    elems += [met, Spacer(1, .4*cm)]

    # Tabla principal
    col_w = [1*cm, 9.5*cm, 2.8*cm, 2.8*cm, 2.8*cm, 3.2*cm, 2.8*cm]
    hdrs  = ['#','SECRETARÍA','EMPLEADOS','YA VOTARON','FALTAN','% PART.','ESTADO']
    hdr_row = [Paragraph(f"<b>{h}</b>",
                ps(f'th{i}', fontSize=8, fontName='Helvetica-Bold',
                   textColor=RL_BLANCO, alignment=TA_CENTER))
               for i, h in enumerate(hdrs)]
    rows = [hdr_row]

    for idx, row in enumerate(data, 1):
        falt = row['empleados'] - row['votos_reportados']
        pct  = (row['votos_reportados'] / row['empleados'] * 100) if row['empleados'] > 0 else 0
        estado = "✔ Completo" if falt==0 else ("▲ Parcial" if pct>0 else "✗ Sin votos")
        pc = colors.HexColor('#27AE60') if pct>=80 else (RL_AMARILLO if pct>=50 else RL_ROJO)
        rows.append([
            Paragraph(str(idx), bold_c),
            Paragraph(row['name'], normal),
            Paragraph(str(row['empleados']), bold_c),
            Paragraph(str(row['votos_reportados']), bold_c),
            Paragraph(str(falt), bold_c),
            Paragraph(f"<b>{pct:.1f}%</b>",
                      ps(f'pct{idx}', fontSize=8, fontName='Helvetica-Bold',
                         textColor=pc, alignment=TA_CENTER)),
            Paragraph(estado, bold_c),
        ])

    rows.append([
        Paragraph("", bold_c),
        Paragraph("<b>TOTAL GENERAL</b>",
                  ps('tot', fontSize=9, fontName='Helvetica-Bold', alignment=TA_LEFT)),
        Paragraph(f"<b>{total_emp}</b>",  bold_c),
        Paragraph(f"<b>{total_vot}</b>",  bold_c),
        Paragraph(f"<b>{total_falt}</b>", bold_c),
        Paragraph(f"<b>{pct_gen:.1f}%</b>", bold_c),
        Paragraph("", bold_c),
    ])

    n = len(rows)
    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),   RL_AZUL),
        ('BACKGROUND',    (0,n-1),(-1,n-1), RL_AMARILLO),
        ('ROWBACKGROUNDS',(0,1),  (-1,n-2), [RL_BLANCO, RL_GRIS]),
        ('GRID',          (0,0),  (-1,-1),  0.4, colors.HexColor('#DDDDDD')),
        ('BOX',           (0,0),  (-1,-1),  1,   RL_AZUL),
        ('VALIGN',        (0,0),  (-1,-1),  'MIDDLE'),
        ('TOPPADDING',    (0,0),  (-1,-1),  4),
        ('BOTTOMPADDING', (0,0),  (-1,-1),  4),
        ('LEFTPADDING',   (0,0),  (-1,-1),  4),
        ('FONTSIZE',      (0,0),  (-1,-1),  8),
    ]))
    elems.append(tbl)

    # Pie
    elems += [
        Spacer(1, .5*cm),
        HRFlowable(width="100%", thickness=2, color=RL_AZUL),
        Spacer(1, .15*cm),
        Paragraph(
            "Consulta Popular Nacional 2026 · Gobernación del Estado Bolívar · Documento generado automáticamente",
            ps('foot', fontSize=7, textColor=colors.HexColor('#888888'), alignment=TA_CENTER)
        )
    ]

    doc.build(elems)
    buf.seek(0)
    fname = f"participacion_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/pdf')
